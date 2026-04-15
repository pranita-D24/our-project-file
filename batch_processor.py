import os
import time
import openpyxl
import re
from openpyxl.styles import Font, PatternFill, Alignment
from comparator import compare
import logging

# Setup Logging
logging.basicConfig(level=logging.INFO, format='%(message)s')
logger = logging.getLogger("BatchProcessor")

def find_pairs_flat(folder):
    """
    Automatically pairs files in a flat folder by matching base filenames 
    with their '- Copy' counterparts.
    """
    files = [f for f in os.listdir(folder) if f.lower().endswith('.pdf')]
    pairs = []
    # Regex to handle various 'Copy' suffix styles: ' - Copy', '-Copy', ' Copy', ' - Copy(1)', etc.
    suffix_pattern = re.compile(r'[-\s]*copy[\d\s\(\)]*', re.IGNORECASE)
    
    for f in files:
        if "copy" in f.lower():
            # Identify the original file matching this copy
            name_only = os.path.splitext(f)[0]
            base_name = suffix_pattern.sub('', name_only).strip()
            
            # Find matching original
            for orig in files:
                orig_name = os.path.splitext(orig)[0].strip()
                if orig_name.lower() == base_name.lower():
                    # Ensure we don't pair a file with itself
                    if orig.lower() != f.lower():
                        # We use (Original, Copy) order for comparison logic
                        # (V1: Original, V2: Modified/Copy)
                        pairs.append((orig, f))
                        break
    return pairs

def run_batch():
    drawings_dir = r"C:\Trivim Internship\engineering_comparison_system\Drawings"
    
    # 1. Dynamically Identify Pairs
    pairs = find_pairs_flat(drawings_dir)
    
    if not pairs:
        print(f"No pairs found in {drawings_dir}")
        return

    print(f"Starting Batch Comparison of {len(pairs)} detected pairs...")
    for orig, copy in pairs:
        print(f"   [PAIR] {orig} <-> {copy}")
    
    master_results = []
    
    for v1_file, v2_file in pairs:
        drawing_id = os.path.splitext(v1_file)[0]
        path1 = os.path.join(drawings_dir, v1_file)
        path2 = os.path.join(drawings_dir, v2_file)
        
        print(f"\n>>> Processing: {drawing_id}")
        try:
            res = compare(path1, path2, drawing_id=drawing_id)
            
            # Use geometry cluster counts from result.geometry
            geo = res.geometry
            stats = {
                "id": drawing_id,
                "verdict": res.verdict,
                "added": len(geo.get("added", [])),
                "removed": len(geo.get("removed", [])),
                "moved": len(geo.get("moved", [])),
                "resized": len(geo.get("resized", [])),
                "dims": len(res.dim_changes) if hasattr(res, 'dim_changes') else 0
            }
            master_results.append(stats)
            print(f"    Done: {res.verdict} | Add: {stats['added']} | Rem: {stats['removed']}")
        except Exception as e:
            print(f"    FAILED: {drawing_id} Error: {e}")
            import traceback; traceback.print_exc()
            
    # 2. Generate Master Summary
    generate_master_summary(master_results, drawings_dir)

def generate_master_summary(results, output_dir):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Master Audit Summary"
    
    headers = ["Drawing ID", "Verdict", "Added", "Removed", "Moved", "Resized", "Dim Changes"]
    ws.append(headers)
    
    # Styling
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        
    for r in results:
        ws.append([
            r["id"], r["verdict"],
            r["added"], r["removed"], r["moved"], r["resized"], r["dims"]
        ])
        
    # Auto-width
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except: pass
        ws.column_dimensions[column].width = max_length + 2
        
    summary_path = os.path.join(output_dir, "Master_Audit_Summary.xlsx")
    wb.save(summary_path)
    print(f"\nBATCH COMPLETE. Master Summary saved to: {summary_path}")

if __name__ == "__main__":
    run_batch()